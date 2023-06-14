from googleapiclient.discovery import build
api_key='AIzaSyCOjfTLdnQOUjGJD_OCNhX0O_6UyIAPG5U'
from googleapiclient.discovery import build
import random as rd
import string
import numpy as np
import seaborn as sns
import pandas as pd
import dash
import matplotlib.pyplot as plt
from dash import Dash, html, dcc,dash_table,callback,Output,Input,State
import plotly.express as px
import openpyxl
import random as rd
import string
import numpy as np
import seaborn as sns
import pandas as pd
import matplotlib.pyplot as plt
# Set up the API client
youtube = build('youtube', 'v3', developerKey=api_key)


class yt_analysis:
    def __init__(self,
                    channelType=None,
                    eventType=None,
                    location=None,
                    locationRadius=None,
                    maxResults=None,
                    channelId=None,
                    regions_code=None,
                    keyword=None,
                    order=None,
                    publishedAfter=None,
                    publishedBefore=None,
                    relevanceLanguage=None):
      
        self.channelType=channelType
        self.eventType=eventType
        self.location=location
        self.locationRadius=locationRadius
        self.maxResults=maxResults
        self.channelId=channelId
        self.regions_code=regions_code
        self.keyword=keyword
        self.order=order
        self.publishedAfter=publishedAfter
        self.publishedBefore=publishedBefore
        self.relevanceLanguage=relevanceLanguage

    def get_keyword(self):
        
        length = rd.randint(1, 8)  # Generate random length between 1 and 8
        letters = string.ascii_letters
        self.keyword =''.join(rd.choice(letters) for _ in range(length))

        return self.keyword 

    # Process the channels and print their ids
        
    def yt_search(self):
        
        if self.maxResults==None:
            self.maxResults=2
            
        if self.keyword==None and self.channelId==None:
            self.keyword=self.get_keyword()  
        
        request = youtube.search().list(
                        channelType=self.channelType,
                        eventType=self.eventType,
                        location=self.location,
                        locationRadius=self.locationRadius,
                        order=self.order,
                        publishedAfter=self.publishedAfter,
                        publishedBefore=self.publishedBefore,
                        regionCode=self.regions_code,
                        relevanceLanguage=self.relevanceLanguage,
                        channelId=self.channelId,
                        part="snippet",
                        q=self.keyword,
                        maxResults=self.maxResults
                        
                    )
        
        channel_response =request.execute()
        channels = channel_response['items']
        self.ids_chan=[]
        for channel in channels: 
            self.ids_chan.append(channel["snippet"]['channelId'])
        return self.ids_chan
 
        
    def chan_stats(self):
            self.ids_chan=self.yt_search()
            stats_request = youtube.channels().list(
                part="snippet,contentDetails,statistics,topicDetails",
                id=','.join(self.ids_chan)
                    )
            chan_response = stats_request.execute()
            self.data=[]
            self.username=[]
            self.subscriber=[]
            self.video_total=[]
            self.view=[]
           
            for i in range(len(chan_response['items'])):
                
                chan= chan_response['items'][i]
                self.usernames=chan['snippet']['title']
                
                
                self.username.append(self.usernames)
                self.subscriberCount=chan['statistics']['subscriberCount']
                
                
                self.subscriber.append(self.subscriberCount)
                self.views=chan['statistics']['viewCount']
                
                self.view.append(self.views)
                self.videos_total=chan['statistics']['videoCount']
                
                self.video_total.append(self.videos_total)
                
                self.data.append(dict(channel_name=self.usernames , subscribers= self.subscriberCount,views= self.views, number_of_videos=self.videos_total))
            return self.data ,  self.username ,self.video_total, self.subscriber , self.view
        
    def viz(self): 
        
        self.data ,self.username ,self.video_total, self.subscriber, self.view =self.chan_stats()
        df=pd.DataFrame(self.data)
        df['subscribers']=pd.to_numeric(df['subscribers'])
        df['views']=pd.to_numeric(df['views'])
        df['number_of_videos']=pd.to_numeric(df['number_of_videos'])
        sns.set(rc={'figure.figsize':(10,10)}) 
        plt.subplot(1,3,1)
        bar_sub=sns.barplot(x='channel_name',y='subscribers',data=df)
        plt.xticks(rotation=45)
        plt.subplot(1,3,2)
        bar_vid=sns.barplot(x='channel_name',y='number_of_videos',data=df)
        plt.xticks(rotation=45)
        plt.subplot(1,3,3)
        bar_views=sns.barplot(x='channel_name',y='views',data=df)
        plt.xticks(rotation=45)
        return  df
       
    def excel(self):
        import openpyxl
        self.data ,self.username ,self.video_total, self.subscriber, self.view =self.chan_stats()
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = "youtube channels report"
        self.sheet_column_names = ["channel name", "number of subscribers", "number of views", "total of videos"]
        for col_num, column_name in enumerate(self.sheet_column_names, start=1):
                self.sheet.cell(row=1, column=col_num).value = column_name
            
         
        for i in range(len(self.username)):
            self.sheet.cell(row=2+i,column=1).value =self.username[i]
            self.sheet.cell(row=2+i,column=2).value =self.subscriber[i]
            self.sheet.cell(row=2+i,column=3).value =self.view[i]
            self.sheet.cell(row=2+i,column=4).value =self.video_total[i]
            
            
            
        self.sheet.sheet_format.defaultColWidth  = 20
        return self.workbook.save("youtube.xlsx")


class dashboard(yt_analysis):
    def __init__(self, channelType=None, eventType=None, location=None, locationRadius=None,
                 maxResults=None, channelId=None, regions_code=None, keyword=None, order=None,
                 publishedAfter=None, publishedBefore=None, relevanceLanguage=None):
        super().__init__(channelType, eventType, location, locationRadius, maxResults, channelId,
                         regions_code, keyword, order, publishedAfter, publishedBefore, relevanceLanguage)

    def run_dashboard(self):
        app = dash.Dash(__name__)
        app.title = "YouTube Channels Analysis"
        self.data, self.username, self.video_total, self.subscriber, self.view = self.chan_stats()
        fig = px.histogram(self.data, x='channel_name', y='subscribers', histfunc='avg')
        app.layout = html.Div(children=[
            html.H1(children='Hello Dash'),
            html.Div(children='''Dash: A web application framework for your data.'''),

            html.Label('Select Inputs to Enter:'),

           dcc.Input(
                id="q",
                value=self.get_keyword(),
                type="text"
            ),

            

          

            html.Button('Update Inputs', id='update-button'),

            html.Button('Save Excel Sheet', id='save-button'),

            dcc.Graph(figure=fig,
                id='example-graph'
            )
        ])

        

        @app.callback(
            Output('example-graph', 'figure'),
            Input('q', "value"),
            Input('update-button', 'n_clicks'),
            
            
        )
        def update_inputs(n_clicks, q):
            if n_clicks is not None:
                self.keyword=q
                self.data, self.username, self.video_total, self.subscriber, self.view = ds.chan_stats()
                fig = px.histogram(self.data, x='channel_name', y='subscribers', histfunc='avg')
                return fig

        @app.callback(
            Output('save-button', 'n_clicks'),
            Input('save-button', 'n_clicks')
        )
        def save_excel_sheet(n_clicks):
            if n_clicks is not None:
                self.excel()
            return None

        app.run_server(port=8052)

ds = dashboard()
ds.run_dashboard()




    
    


    











             
         