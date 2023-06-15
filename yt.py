from  lib_api import *

class yt_analysis:
    def __init__(self,
                    channelType=None,
                    eventType=None,
                    location=None,
                    locationRadius=None,
                    maxResults=None,
                    channelId=None,
                    regions_code=None,
                    keyword=None,
                    order=None,
                    publishedAfter=None,
                    publishedBefore=None,
                    relevanceLanguage=None):
      
        self.channelType=channelType
        self.eventType=eventType
        self.location=location
        self.locationRadius=locationRadius
        self.maxResults=maxResults
        self.channelId=channelId
        self.regions_code=regions_code
        self.keyword=keyword
        self.order=order
        self.publishedAfter=publishedAfter
        self.publishedBefore=publishedBefore
        self.relevanceLanguage=relevanceLanguage

    def get_keyword(self):
        
        length = rd.randint(1, 8)  # Generate random length between 1 and 8
        letters = string.ascii_letters
        self.keyword =''.join(rd.choice(letters) for _ in range(length))

        return self.keyword

    # Process the channels and print their ids
        
    def yt_search(self):
        
        if self.maxResults==None:
            self.maxResults=5
     
        if self.keyword==None and self.channelId==None:
            self.keyword=self.get_keyword() 
      
        
        request = youtube.search().list(
                        channelType=self.channelType,
                        eventType=self.eventType,
                        location=self.location,
                        locationRadius=self.locationRadius,
                        order=self.order,
                        publishedAfter=self.publishedAfter,
                        publishedBefore=self.publishedBefore,
                        regionCode=self.regions_code,
                        relevanceLanguage=self.relevanceLanguage,
                        channelId=self.channelId,
                        part="snippet",
                        q=self.keyword,
                        maxResults=self.maxResults
                        
                    )
        
        channel_response =request.execute()
        channels = channel_response['items']
        self.ids_chan=[]
        for channel in channels: 
            self.ids_chan.append(channel["snippet"]['channelId'])
        return self.ids_chan
 
        
    def chan_stats(self):
            self.ids_chan=self.yt_search()
            stats_request = youtube.channels().list(
                part="snippet,contentDetails,statistics,topicDetails",
                id=','.join(self.ids_chan)
                    )
            chan_response = stats_request.execute()
            self.data=[]
            self.username=[]
            self.subscriber=[]
            self.video_total=[]
            self.view=[]
           
            for i in range(len(chan_response['items'])):
                
                chan= chan_response['items'][i]
                self.usernames=chan['snippet']['title']
                
                
                self.username.append(self.usernames)
                self.subscriberCount=chan['statistics']['subscriberCount']
                
                
                self.subscriber.append(self.subscriberCount)
                self.views=chan['statistics']['viewCount']
                
                self.view.append(self.views)
                self.videos_total=chan['statistics']['videoCount']
                
                self.video_total.append(self.videos_total)
                
                self.data.append(dict(channel_name=self.usernames , subscribers= self.subscriberCount,views= self.views, number_of_videos=self.videos_total))
            return self.data ,  self.username ,self.video_total, self.subscriber , self.view
        
    def viz(self): 
        
        self.data ,self.username ,self.video_total, self.subscriber, self.view =self.chan_stats()
        df=pd.DataFrame(self.data)
        df['subscribers']=pd.to_numeric(df['subscribers'])
        df['views']=pd.to_numeric(df['views'])
        df['number_of_videos']=pd.to_numeric(df['number_of_videos'])
        sns.set(rc={'figure.figsize':(10,10)}) 
        plt.subplot(1,3,1)
        bar_sub=sns.barplot(x='channel_name',y='subscribers',data=df)
        plt.xticks(rotation=45)
        plt.subplot(1,3,2)
        bar_vid=sns.barplot(x='channel_name',y='number_of_videos',data=df)
        plt.xticks(rotation=45)
        plt.subplot(1,3,3)
        bar_views=sns.barplot(x='channel_name',y='views',data=df)
        plt.xticks(rotation=45)
        return  df
       
    def excel(self):
        import openpyxl
        self.data ,self.username ,self.video_total, self.subscriber, self.view =self.chan_stats()
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = "youtube channels report"
        self.sheet_column_names = ["channel name", "number of subscribers", "number of views", "total of videos"]
        for col_num, column_name in enumerate(self.sheet_column_names, start=1):
                self.sheet.cell(row=1, column=col_num).value = column_name
            
         
        for i in range(len(self.username)):
            self.sheet.cell(row=2+i,column=1).value =self.username[i]
            self.sheet.cell(row=2+i,column=2).value =self.subscriber[i]
            self.sheet.cell(row=2+i,column=3).value =self.view[i]
            self.sheet.cell(row=2+i,column=4).value =self.video_total[i]
            
            
            
        self.sheet.sheet_format.defaultColWidth  = 20
        return self.workbook.save("youtube_channels.xlsx")





    
    


    











             
         